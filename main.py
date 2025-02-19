import os

from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem
from PyQt6 import uic
import sys
import openpyxl
from data.students import Students
from data.subjects import Subjects

from data import db_session

SCREEN_SIZE_X, SCREEN_SIZE_Y = (847, 709)  # Окно импорта
SCREEN_SIZE_X1, SCREEN_SIZE_Y1 = (1311, 730)  # Окно информации из базы данных

SUB = {'Матем': 'Математика',
       'Физика': 'Физика',
       'Химия': 'Химия',
       'История': 'История',
       'Рус яз': 'Русский язык',
       'Биолог': 'Биология',
       'Литра': 'Литература',
       'Геогр': 'География',
       'Инфо': 'Информатика',
       'Обществ': 'Обществознание',
       'Англ яз': 'Английский язык',
       'Франц яз': 'Французский язык',
       'Немец яз': 'Немецкий язык',
       'Испан яз': 'Испанский язык'}


def main():
    try:
        os.remove("db/exeldata.db")
    except OSError:
        pass
    db_session.global_init("db/exeldata.db")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('designs/projectdesign2224.ui', self)

        self.setFixedSize(SCREEN_SIZE_X, SCREEN_SIZE_Y)
        self.import_btn.clicked.connect(self.import_xlsx)

    def sql_exel(self, list_students):
        db_sess = db_session.create_session()
        for stud in list_students:
            new_student = Students(name=stud[0],
                                   number=stud[1],
                                   parent1=stud[2],
                                   parent2=stud[3],
                                   subj1=stud[4],
                                   subj2=stud[5],
                                   subj3=stud[6],
                                   subj4=stud[7],
                                   subj5=stud[8],
                                   subj6=stud[9],
                                   subj7=stud[10],
                                   subj8=stud[11],
                                   subj9=stud[12],
                                   subj10=stud[13],
                                   subj11=stud[14],
                                   subj12=stud[15],
                                   subj13=stud[16])
            db_sess.add(new_student)
            db_sess.commit()

    def import_xlsx(self):
        global dataWindow
        fname = QFileDialog.getOpenFileName(self, 'Выберите файл', None,
                                            'Файл (*.xlsx)')[0]
        if fname:
            workbook = openpyxl.load_workbook(fname)
            sheet = workbook.active  # Получение активного листа
            rows = sheet.max_row
            list_exel = []
            lesson = []
            for elem in sheet['A1':'R1']:
                for el in elem:
                    lesson.append(el.value)  # Запоминание всех предметов
            for row in sheet.iter_rows(min_row=2, max_row=rows, values_only=True):
                list_exel.append(row)  # сохраняет в список значения ячеек с 1 по последнюю строки
            print(list_exel)
            print(lesson)
            db_sess = db_session.create_session()
            subject = Subjects()
            for el in lesson[4:]:
                subject.title = SUB[el]
                db_sess.merge(subject)
            db_sess.commit()
            workbook.close()
            dataWindow = TableDataBase(list_exel, lesson)
            dataWindow.show()
        self.sql_exel(list_exel)


class TableDataBase(MainWindow):
    def __init__(self, students, subjects):
        super().__init__()
        uic.loadUi('designs/databaseStudents.ui', self)
        self.students = students
        self.subjects = subjects
        self.setFixedSize(SCREEN_SIZE_X1, SCREEN_SIZE_Y1)
        self.tableWidget.setRowCount(len(self.students))
        print(self.students)
        for stud in range(len(self.students)):
            for data in range(len(self.students[stud])):
                if self.students[stud][data]:
                    self.tableWidget.setItem(stud, data, QTableWidgetItem(str(self.students[stud][data])))
                else:
                    self.tableWidget.setItem(stud, data, QTableWidgetItem('Пусто'))



if __name__ == '__main__':
    main()
    app = QApplication(sys.argv)
    ex = MainWindow()
    ex.show()
    sys.exit(app.exec())
